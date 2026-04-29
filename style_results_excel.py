"""Style CodonDomesticate result CSV files as Excel workbooks."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _column_width(header: str, values: list[str]) -> float:
    if header in {
        "sequence",
        "input_cds",
        "domesticated_cds",
        "final_cds",
        "protein_before",
        "protein_after",
    }:
        return 42
    if header in {
        "source_file",
        "site_hits_before",
        "mutations",
        "aa_change_mutations",
        "aa_change_codon_usage_status",
        "error",
    }:
        return 36
    longest = max([len(header), *(min(len(str(value)), 60) for value in values)], default=len(header))
    return min(max(longest + 3, 10), 28)


def style_results_csv_as_xlsx(
    csv_path: str | Path,
    xlsx_path: str | Path | None = None,
    sheet_name: str = "domesticated_cds_results",
) -> str:
    """Create a styled XLSX from a CSV without changing any cell values."""
    csv_path = Path(csv_path)
    if xlsx_path is None:
        xlsx_path = csv_path.with_suffix(".xlsx") if csv_path.suffix else Path(str(csv_path) + ".xlsx")
    else:
        xlsx_path = Path(xlsx_path)

    with csv_path.open(newline="", encoding="utf-8") as handle:
        rows = [[cell for cell in row] for row in csv.reader(handle)]
    if not rows:
        raise ValueError(f"CSV has no rows: {csv_path}")
    if len({len(row) for row in rows}) != 1:
        raise ValueError(f"CSV rows have inconsistent column counts: {csv_path}")

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    for row in rows:
        ws.append(row)

    max_row = len(rows)
    max_col = len(rows[0])
    last_col = get_column_letter(max_col)
    headers = rows[0]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    odd_fill = PatternFill("solid", fgColor="FFFFFF")
    even_fill = PatternFill("solid", fgColor="EAF3F8")
    true_fill = PatternFill("solid", fgColor="D9EAD3")
    false_fill = PatternFill("solid", fgColor="F4CCCC")
    strategy_fills = {
        "third_D_to_V": PatternFill("solid", fgColor="FFF2CC"),
        "second_H_to_A_when_third_not_D": PatternFill("solid", fgColor="D9D2E9"),
    }
    border_side = Side(style="thin", color="B7C9D6")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{last_col}{max_row}"
    ws.sheet_view.showGridLines = False

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    strategy_col = headers.index("mhd_mutation_strategy") + 1 if "mhd_mutation_strategy" in headers else None
    for row_idx in range(2, max_row + 1):
        base_fill = even_fill if row_idx % 2 == 0 else odd_fill
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row_idx, col_idx)
            cell.fill = base_fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)
            cell.number_format = "@"
            if cell.value == "True":
                cell.fill = true_fill
            elif cell.value == "False":
                cell.fill = false_fill
        if strategy_col is not None:
            strategy_cell = ws.cell(row_idx, strategy_col)
            strategy_cell.fill = strategy_fills.get(strategy_cell.value, strategy_cell.fill)

    for col_idx in range(1, max_col + 1):
        header = headers[col_idx - 1]
        values = [row[col_idx - 1] for row in rows[1:]]
        ws.column_dimensions[get_column_letter(col_idx)].width = _column_width(header, values)

    ws.row_dimensions[1].height = 42
    for row_idx in range(2, max_row + 1):
        ws.row_dimensions[row_idx].height = 22

    wb.save(xlsx_path)
    _assert_xlsx_values_match(rows, xlsx_path, ws.title)
    return str(xlsx_path)


def _assert_xlsx_values_match(rows: list[list[str]], xlsx_path: Path, sheet_name: str) -> None:
    check_wb = load_workbook(xlsx_path, data_only=False, read_only=True)
    check_ws = check_wb[sheet_name]
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, expected in enumerate(row, start=1):
            actual = check_ws.cell(row_idx, col_idx).value
            actual_text = "" if actual is None else str(actual)
            if actual_text != expected:
                raise AssertionError(f"XLSX value mismatch at row {row_idx}, column {col_idx}")
